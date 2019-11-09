from django.core.management.base import BaseCommand, CommandError
from django.db.models import Count,Sum
# from mach.models import MachTask
from case002.models import Data2

import os
import array
import openpyxl
import statistics as stats

from openpyxl import load_workbook # to read

     # https://openpyxl.readthedocs.io/en/stable/tutorial.html#accessing-one-cell
from openpyxl import Workbook # to write

# 2019-6-15
# 為什麼發現 扣除未流轉，沒能更新到表單？
class Command(BaseCommand):
    help = 'Update Zone19 data'

    # def add_arguments(self, parser):
    #     parser.add_argument('src', nargs='+', type=str)

    def genCat(srcSheetName,cat,idStart,idDetStart,outFileName):

        from openpyxl import load_workbook # to read

        # https://openpyxl.readthedocs.io/en/stable/tutorial.html#accessing-one-cell
        from openpyxl import Workbook # to write


        wb = load_workbook(filename='contract_src.xlsx', read_only=True)

        wb1 = Workbook()
        ws1 = wb1.active

        # ws1.title ="杏福农业主檔"
        ws1.title =cat+"主檔"

        ws1.cell(1,1,'id')
        ws1.cell(1,2,'cat')
        ws1.cell(1,3,'name')
        ws1.cell(1,4,'code')
        ws1.cell(1,5,'sign')
        ws1.cell(1,6,'doer')

        wb2 = Workbook()
        ws2 = wb2.active
        # ws2.title ="杏福农业分項"
        ws2.title =cat+"分項"
        ws2.cell(1,1,'id')
        ws2.cell(1,2,'contract')
        ws2.cell(1,3,'det')
        ws2.cell(1,4,'content')
        ws2.cell(1,5,'price')
        ws2.cell(1,6,'area')
        ws2.cell(1,7,'amt')
        ws2.cell(1,8,'style')



        # cat = '杏福园工程'
        # ws = wb['杏福农业']
        ws = wb[srcSheetName]


        # id = 0
        id = idStart - 1

        # idDet = 0
        idDet = idDetStart - 1

        det = 0
        rec = 0
        for row in ws.rows:
            rec += 1
            if rec < 3:
                continue
            # if rec > 57:
            #     continue

            colA = row[0].value
            colB = row[1].value
            colC = row[2].value
            colD = row[3].value
            colE = row[4].value
            colF = row[5].value
            colG = row[6].value
            colH = row[7].value
            colI = row[8].value

            if colA is None and colB is None and colC is None:
                continue

            if colA:
                id += 1
                idDet += 1

                det = 1
                fixA=colA
                fixB=colB
                fixG=colG
                fixH=colH
                fixI=colI

                # print (rec,id,det,cat1,fixA,fixB,colC,colD,colE,colG,fixH,fixI)
                print (rec,id," ",cat,fixA,fixB,fixH,fixI)
                print (rec,id,det,colC,fixG)
                # ws1.cell(row=1+id, column=1, value=id)
                rowNum = 1+id -idStart + 1
                ws1.cell(rowNum, 1, id)
                ws1.cell(rowNum, 2, cat)
                ws1.cell(rowNum, 3, fixA)
                ws1.cell(rowNum, 4, fixB)
                ws1.cell(rowNum, 5, fixH)
                ws1.cell(rowNum, 6, fixI)

            else:
                idDet += 1
                det += 1
                # print (rec,id,det,cat1,fixA,fixB,fixH,fixI)
                print (rec,id,det,colC,fixG)


            rowNum = 1+idDet -idDetStart + 1
            ws2.cell(rowNum,1,idDet)
            ws2.cell(rowNum,2,id)
            ws2.cell(rowNum,3,det)
            ws2.cell(rowNum,4,colC)
            ws2.cell(rowNum,5,colD)
            ws2.cell(rowNum,6,colE)
            ws2.cell(rowNum,7,colF)
            ws2.cell(rowNum,8,fixG)


            # for cell in row:
            #     print(cell.value)
        wb1.save(outFileName+cat+".xlsx")
        # wb2.save("contact_out_2.xlsx")
        wb2.save(outFileName+cat+"det.xlsx")
    # genCat('杏福农业','杏福园工程',1,1,'out')
    # genCat('秉承物流','秉承物流',50,54,'out')

        
    def handle(self, *args, **options):
        def get_sheet_data(filename,sheetname):
            wb = load_workbook(filename=src,read_only=False)
            ws = wb[sheetname]
            def m1(ws):
                def get_val(cell):
                    if row[1].value is None:
                        return 'xxx'
                    else:
                        return str(cell.value)
               
                for row in ws.rows:
                    A= row[0].value
                    # self.stdout.write(A)
                    B = get_val(row[1])
                    C = get_val(row[2])
                    # C = 'c'
                    
                    D = get_val(row[3])
                    E = get_val(row[4])
                    F = get_val(row[5])
                    G = get_val(row[6])
                    H = get_val(row[7])
                    I = get_val(row[8])
                    J = get_val(row[9])
                    K = get_val(row[10])
                    # self.stdout.write(A)
                    # self.stdout.write(B)
                    # self.stdout.write(C)
                    # self.stdout.write(D)
                    # self.stdout.write(E)
                    # self.stdout.write(F)
                    # self.stdout.write(G)
                    # self.stdout.write(H)
                    # self.stdout.write(I)
                    # self.stdout.write(J)
                    # self.stdout.write(K)
                    temp = A+' '+ B
                    self.stdout.write(temp)
                    
               
            m1(ws)
    
        oscwd =os.getcwd
        cwd='/Users/pinglingchen/django-team/jungle123/case002/case002/management/commands/'

        src = cwd+'src.xlsx'
        self.stdout.write(src)
        # self.stdout.write(self.style.SUCCESS('Current Wordking Directory is "%s"' % cwd))
        
        self.stdout.write(self.style.SUCCESS('This program is "%s"' % 'case002-import.py'))
        
        # https://openpyxl.readthedocs.io/en/stable/optimized.html
        
        wb = load_workbook(filename=src,read_only=False)
        # ws = wb['6.4']
        
        

        def m1(ws):

            def get_val(cell):
                if row[1].value is None:
                    return 'xxx'
                else:
                    return str(cell.value)
        
            for row in ws.rows:
                A= row[0].value
                # self.stdout.write(A)
                B = get_val(row[1])
                C = get_val(row[2])
                # C = 'c'
                
                D = get_val(row[3])
                E = get_val(row[4])
                F = get_val(row[5])
                G = get_val(row[6])
                H = get_val(row[7])
                I = get_val(row[8])
                J = get_val(row[9])
                K = get_val(row[10])
                self.stdout.write(A)
                self.stdout.write(B)
                self.stdout.write(C)
                self.stdout.write(D)
                self.stdout.write(E)
                self.stdout.write(F)
                self.stdout.write(G)
                self.stdout.write(H)
                self.stdout.write(I)
                self.stdout.write(J)
                self.stdout.write(K)
                             
                
              
        # m1(ws)

        def m2(sheetname):
            ws =wb[sheetname]
            
            def get_val(cell):
                if row[1].value is None:
                    return 'xxx'
                else:
                    return str(cell.value)
            rec =0
            for row in ws.rows:
                A= get_val(row[0])
                    # return
                # self.stdout.write(A)
                B = get_val(row[1])
                C = get_val(row[2])
                # C = 'c'
                
                D = get_val(row[3])
                E = get_val(row[4])
                F = get_val(row[5])
                G = get_val(row[6])
                H = get_val(row[7])
                I = get_val(row[8])
                J = get_val(row[9])
                K = get_val(row[10])

                if A == 'xxx':
                    pass
                else:
              
                    rec += 1
                    self.stdout.write('****')
                    self.stdout.write(str(rec))
                    self.stdout.write('****')
                    self.stdout.write(A)
                    self.stdout.write(B)
                    self.stdout.write(C)
                    self.stdout.write(D)
                    self.stdout.write(E)
                    self.stdout.write(F)
                    self.stdout.write(G)
                    self.stdout.write(H)
                    self.stdout.write(I)
                    self.stdout.write(J)
                    self.stdout.write(K)
                                
                
              
        # m2('6.1')
        # m2('6.2')
        # m2('6.3')

        def m3(src,sheetname):
            wb = load_workbook(filename=src,read_only=False)
        
            ws =wb[sheetname]
            
            def get_val(cell):
                if cell == 'None':
                    return '---'
                if row[1].value in [None,'None']:
                    return 'xxx'
                else:
                    if str(cell.value) == 'None':
                        return '---'
                    
                    return str(cell.value)
            rec = 0
            mem_cnt = 0
            for row in ws.rows:
                rec += 1
                    
                A = get_val(row[0])
                # self.stdout.write(A)
                B = get_val(row[1])
                #  4 if var1 is None else var1
                C = get_val(row[2])
                
                D = get_val(row[3])
                E = get_val(row[4])
                F = get_val(row[5])
                G = get_val(row[6])
                H = get_val(row[7])
                I = get_val(row[8])
                J = get_val(row[9])
                K = get_val(row[10])
               
                if B == 'Member':
                    mem_cnt += 1
              
                    # self.stdout.write('****')
                    # self.stdout.write(str(rec))
                    # self.stdout.write('****')
                    # self.stdout.write(A)
                    # self.stdout.write(B)
                    # self.stdout.write(C)
                    # self.stdout.write(D)
                    # self.stdout.write(E)
                    # self.stdout.write(F)
                    # self.stdout.write(G)
                    # self.stdout.write(H)
                    # self.stdout.write(I)
                    # self.stdout.write(J)
                    # self.stdout.write(K)
                
                    temp = str(rec)+' '+str(mem_cnt)+' '+A+' '+ B+' '+ C+' '+ E+' '+ G
                    # temp = str(rec)+' '+str(mem_cnt)+' '+A+' '+ B+' '+ C
                    self.stdout.write(temp)
                    # obj = Data2(name=A,member='Member',date1='2019-06-05',role=C)
                    # obj.save()


                    def addData2(name,member, date1, role):
                        obj = Data2(name=name,member=member,date1=date1,role=role)
                        obj.save()
                    addData2(A,'Member','2019-06-05',C)
                    addData2(A,'Member','2019-06-12',E)
                    addData2(A,'Member','2019-06-19',G)
                    addData2(A,'Member','2019-06-26',I)

                    addData2(A,'Member','2019-07-03',get_val(row[11]))
                    addData2(A,'Member','2019-07-10',get_val(row[13]))
                    addData2(A,'Member','2019-07-17',get_val(row[15]))
                    addData2(A,'Member','2019-07-24',get_val(row[17]))
                    addData2(A,'Member','2019-07-31',get_val(row[19]))

                    addData2(A,'Member','2019-08-07',get_val(row[22]))
                    addData2(A,'Member','2019-08-14',get_val(row[24]))
                    addData2(A,'Member','2019-08-21',get_val(row[26]))
                    addData2(A,'Member','2019-08-28',get_val(row[28]))
                    
                    addData2(A,'Member','2019-09-04',get_val(row[31]))
                    addData2(A,'Member','2019-09-11',get_val(row[33]))
                    addData2(A,'Member','2019-09-18',get_val(row[35]))
                    addData2(A,'Member','2019-09-25',get_val(row[37]))
                    
                    addData2(A,'Member','2019-10-09',get_val(row[39]))
                    addData2(A,'Member','2019-10-16',get_val(row[41]))
                    addData2(A,'Member','2019-10-23',get_val(row[43]))
                    # addData2(A,'Member','2019-08-28',get_val(row[28]))

                    
        Data2.objects.filter(member='Member').delete()
        m3(src,'Registration')
                # Registration
        def fixPoints():
            list1 = Data2.objects.all()
            for x in list1:
                if x.role == 'Speaker':
                    x.points= 3
                    x.save()

                if x.role == 'Attendance':
                    x.points= 1
                    x.save()

                if x.role in ['Ah-counter','GE','Grammarian','IE','TME','TT Evaluator','TT-master','Timer'] :
                    x.points= 2
                    x.save()

                
                    
        fixPoints()